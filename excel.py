# pip install openpyxl
import openpyxl
import os
from openpyxl import Workbook
import time
from openpyxl import styles


dir(openpyxl) #open the dir
wb  = Workbook()   #wb = workbook
sheet = wb.active #new sheet make comment
sheet
sheet["A1"] = "name" #what ever u enter etc...
sheet["B1"] = "age" #what ever u enter etc...
wb.save(filename = "test.xlsx") #file name enter i am enter test.xlsx u enter etc...
os.getcwd() #find the dir comment
wb.save(filename = "C:\\download\\test.xlsx") # dir path name c diver /download folder
        
# step 2 sheet update comment
sheet1 =wb.create_sheet() # without name
sheet2 =wb.create_sheet("test book") # with name
sheet3  =wb.create_sheet("test1 book") # with name
sheet4  =wb.create_sheet("test2 book",0) # with name with opstion
sheet5  =wb.create_sheet("test2 book",1) # with name with opstion

# step 3 sheet Rename comment
sheet1.title = "New Title" # rename replese old name

# step 4 colour change comment
sheet3.sheet_properties.tabColor = "FF0000" # only for colour code 

# step 5 all type 
sheet["A1"] = 100
sheet["A2"] = 10.123
sheet["A3"] = "welcome"
import time
now = time.strftime("%x") # time format
sheet["A4"] = now

v1 = sheet["B1"] # B1 shell now
v1.value = 200
sheet.cell(row = 3, column = 3).value = "2000"

# step 6 styles 
from openpyxl import styles
s = wb.active
s
s["B4"] = "Santra"
help(styles.Font)
help(Workbook.sheet)

s["B4"].font = styles.Font(name = "Bauhaus 93", size = 22, bold = True, italic = True, vertAlign = None, underline = None, strike = True, rgb = "000000FF")

# step 7 read the excel files
path = "D:\\VS CODE\\Excel Python\\python_excel\text.xlsx" # path select 
wb_file = openpyxl.load_workbook(path)
we_file
sheet = wb_file.active
c3 = sheet.cell(row = 3, column = 3)
c3
c3.value

# step 8 how many row avali in excel
sheet.max_row #up to down
sheet.max_column # rigth to left
sheet

# this one column conut
max_col = sheet.max_column
max_col
for i in range(1, max_col + 1):
    cell_vel =sheet.cell(row = 1, column = i)
    print(cell_vel.value)

# this one row count
max_rows = sheet.max_row
max_rows
for i in range(1, max_rows + 1):
    cell_vel = sheet.cell(column = 1, row = i)
    print(cell_vel.value)

# step 9 margin and unmagin
sheet = wb.active
sheet.merge_cells("A1:C1") # merge
wb.save("text.xlsx")
sheet.unmerge_cells("A1:C1") # unmerge

# step 10 size row and column
wb = Workbook()
sheet = wb.active
sheet["A1"] = "test"
sheet.row_dimensions[1].height = 70
sheet.column_dimensions["A"].width = 20
wb.save ("text.xlsx")
wb.close
